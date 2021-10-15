# coding=utf-8
import openpyxl, os, datetime, re, xlrd, copy
#import win32com.client as win32
from openpyxl.workbook import Workbook
import numpy as np
import itertools
from itertools import groupby
from collections import Counter
from itertools import chain, zip_longest




time_schedule = {'08:00 - 09:30':1,'09:40 - 11:10':2,'11:30 - 13:00':3,'13:10 - 14:40':4,'15:00 - 16:30':5,'16:40 - 18:10':6,'18:20 - 19:50':7}
time_schedule1 = {value:key for key, value in time_schedule.items()}
weekday_list = {"Понедельник":1,"Вторник":2,"Среда":3,"Четверг":4,"Пятница":5,"Суббота":6,"Воскресенье":7}
weekday_list1 = {value:key for key, value in weekday_list.items()}

a_dict = {"A":1, "B":2, "C":3, "D":4, "E":5, "F":6, "G":7, "H":8, "I":9, "J":10, "K":11, "L":12, "M":13, "N":14, "O":15, "P":16, "Q":17, "R":18, "S":19, "T":20, "U":21, "V":22, "W":23, "X":24, "Y":25, "Z":26, "AA":27, "AB":28, "AC":29, "AD":30, "AE":31, "AF":32}
#a_dict = list(map(chr, range(97,123))) #для поиска группы по столбцам
a_dict1 = {value:key for key, value in a_dict.items()}
a = range(4, 7) #для поиска группы по строкам
weekday_range = range(4, 100) #для поиска дня недели
length_list = (3, 5) # сколько может быть пар в один день  с учетом пустых ячеек




#liva_weekday = input('Введите день недели:')

def open_xls_as_xlsx(filename):
    # first open using xlrd
    book = xlrd.open_workbook( filename =filename)
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.active

    for row in range(0, nrows):
        for col in range(0, ncols):
            sheet1.cell(row=row+1, column=col+1).value = sheet.cell_value(row, col)

    return book1

def read_weekday(last_time, weekday_parse):
    to_day_or_not_today = ''

    #текущее время
    offset = datetime.timezone(datetime.timedelta(hours=3))
    now = datetime.datetime.now(offset)
    now = now.strftime("%H:%M")

    print('Текущее время:', now, 'Время окончания пары:', last_time)

    if weekday_parse == 'нет':
        if weekday_list1.get(datetime.datetime.today().isoweekday()) != 'Воскресенье':
            # получаем день недели
            if now > last_time: #если последняя пара закончилась, то к дню недели прибовляем +1. Чтобы показать расписание на следующий день
                to_day_or_not_today = ' завтра'
                liva_weekday = weekday_list1.get(datetime.datetime.today().isoweekday()+1)
                if liva_weekday.find("Воскресенье") != -1:
                    liva_weekday = 'Понедельник'
                    to_day_or_not_today = ' Понедельник'
            elif now > last_time:
                liva_weekday = weekday_list1.get(datetime.datetime.today().isoweekday() + 1)
                to_day_or_not_today = ' ' + str(liva_weekday)
            else:
                to_day_or_not_today = ' сегодня'
                liva_weekday = weekday_list1.get(datetime.datetime.today().isoweekday())

            if ' Воскресенье' in liva_weekday:
                liva_weekday = ' Понедельник'
        else:
            to_day_or_not_today = ' завтра'
            liva_weekday = weekday_list1.get(datetime.datetime.today().isoweekday()-6)


        #liva_weekday = 'Пятница'

        if to_day_or_not_today.replace(' ', '') == 'Воскресенье':
            to_day_or_not_today = ' Понедельник'
    else:
        if weekday_parse == 'пн':
            liva_weekday = ' Понедельник'
        elif weekday_parse == 'вт':
            liva_weekday = ' Вторник'
        elif weekday_parse == 'ср':
            liva_weekday = ' Среда'
        elif weekday_parse == 'чт':
            liva_weekday = ' Четверг'
        elif weekday_parse == 'пт':
            liva_weekday = ' Пятница'
        elif weekday_parse == 'сб':
            liva_weekday = ' Суббота'
    return liva_weekday




def main_open(group, file_name, function, next_day, merge, weekday_parse):
    if not os.path.exists(f"{file_name}.xlsx"):   #если этого файла нет, создаем новый
        open_xls_as_xlsx(os.getcwd() + f'/{file_name}.xls').save(filename = f'{file_name}.xlsx')
        """
        fname = (os.getcwd() + f"\\{file_name}.xls").replace('\\', '\\') #ищем файл со старым расширением
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(fname)
        wb.SaveAs(fname+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension              сохраняем в новом формате
        wb.Close()                               #FileFormat = 56 is for .xls extension
        excel.Application.Quit()
        """


    # чекаем файл
    wb = openpyxl.reader.excel.load_workbook(filename=f"{file_name}.xlsx")
    wb.active = 0
    sheet = wb.active

    last_time = find_last_time(group, sheet)
    print('[Main open] Файл', file_name, ' группа:', group, 'Последнее время:', last_time, 'Лист:', sheet.title)
    if type(last_time) != type(None):
        if function == 'Расписание':
            return read_file(group, sheet, last_time, next_day, merge, weekday_parse)
        elif function == 'Узнать на какой день недели парсить замены':
            otvet = read_weekday(last_time, weekday_parse)
            print(otvet)
            return otvet
    else:
        return None






def read_file(group, sheet, last_time, next_day, merge, weekday_parse):


    to_day_or_not_today = ''

    #текущее время
    offset = datetime.timezone(datetime.timedelta(hours=3))
    now = datetime.datetime.now(offset)
    now = now.strftime("%H:%M")

    print('Текущее время:', now, 'Время окончания пары:', last_time)

    next_parse_day = 0
    if next_day == True:
        next_parse_day += 1

    if weekday_parse == 'нет':
        if weekday_list1.get(datetime.datetime.today().isoweekday() + next_parse_day) != 'Воскресенье':
            # получаем день недели
            if now > last_time and next_day == False: #если последняя пара закончилась, то к дню недели прибовляем +1. Чтобы показать расписание на следующий день
                to_day_or_not_today = ' завтра'
                liva_weekday = weekday_list1.get(datetime.datetime.today().isoweekday()+1)
                if liva_weekday.find("Воскресенье") != -1:
                    liva_weekday = 'Понедельник'
                    to_day_or_not_today = ' Понедельник'
                print('Пары закончились, идет сравнение с расписанием на следущий день',liva_weekday)
            elif now > last_time and next_day == True:
                liva_weekday = weekday_list1.get(datetime.datetime.today().isoweekday() + next_parse_day + 1)
                to_day_or_not_today = ' ' + str(liva_weekday)
            elif next_day == True:
                liva_weekday = weekday_list1.get(datetime.datetime.today().isoweekday() + next_parse_day)
                to_day_or_not_today = ' ' + str(liva_weekday)
            else:
                to_day_or_not_today = ' сегодня'
                liva_weekday = weekday_list1.get(datetime.datetime.today().isoweekday())
            print('[TEST WEEKDAY] первый выбор')

            if ' Воскресенье' in liva_weekday:
                liva_weekday = ' Понедельник'
        else:
            print("Сегодня воскресенье")
            to_day_or_not_today = ' завтра'
            liva_weekday = weekday_list1.get(datetime.datetime.today().isoweekday()-6)
    else:
        if weekday_parse == 'пн':
            to_day_or_not_today = ' Понедельник'
            liva_weekday = 'Понедельник'
        elif weekday_parse == 'вт':
            to_day_or_not_today = ' Вторник'
            liva_weekday = 'Вторник'
        elif weekday_parse == 'ср':
            to_day_or_not_today = ' Среда'
            liva_weekday = 'Среда'
        elif weekday_parse == 'чт':
            to_day_or_not_today = ' Четверг'
            liva_weekday = 'Четверг'
        elif weekday_parse == 'пт':
            to_day_or_not_today = ' Пятница'
            liva_weekday = 'Пятница'
        elif weekday_parse == 'сб':
            to_day_or_not_today = ' Суббота'
            liva_weekday = 'Суббота'

    #liva_weekday = 'Пятница'
    print('Будет показано расписание на: ', liva_weekday)

    if to_day_or_not_today.replace(' ', '') == 'Воскресенье':
        to_day_or_not_today = ' Понедельник'

    weekly_schedule = []
    for number in a: #цифры/строки
        for key in a_dict: #буквы/столбцы
            if type(sheet[key + str(number)].value) != type(None): #если ячейка не пустая
                if group.upper() in sheet[key + str(number)].value: #ищем группу в расписании
                    #пробегаем в цикле по ячейки для поиска дня недели
                    for day in weekday_list: # пробигаемся по дням неделям
                        if day == liva_weekday: # если мы нашли день недели равным текущему
                            for number_day in weekday_range: # пробигаемся по определенном диапозоне для поиска ...
                                #if type(sheet[key + str(number_day)].value) != type(None) or type(sheet['A' + str(number_day)].value) != type(None): #если ячейка не пустая
                                if type(sheet['A' + str(number_day)].value) != type(None): #если ячейка не пустая
                                
                                    if (sheet['A' + str(number_day)].value).find(day.upper()) != -1: #Нашли день недели
                                        print('[FIND] Ячейка:', sheet['A' + str(number_day)].value)
                                        next_day = weekday_list.get(day)+1 # поменчаем, что неделя следующая (если сейчас 5, то парсим до 6)
                                        lest_day_text = weekday_list1.get(next_day)# ищем номер недели в списке (чтобы писалась неделя текстом, а не цифрой)
                                        
                                        next_weekday_range = range(number_day, 100) # Диапозон от сегодняшнего дня недели до не найденного следующего (будем искать от сегодняшней до 100 ячейки)
                                        
                                        for day_next in weekday_list:
                                            if day_next == lest_day_text:
                                                for number_day_next in next_weekday_range: #пробегаем в цикле по ячейки для поиска след дня недели
                                                    #if type(sheet[key + str(number_day_next)].value) != type(None) or type(sheet['A' + str(number_day_next)].value) != type(None): #если ячейка не пустая
                                                    if type(sheet['A' + str(number_day_next)].value) != type(None): #если ячейка не пустая    
                                                        if (sheet['A' + str(number_day_next)].value).find(day_next.upper()) != -1 or day_next == 'Воскресенье': #Нашли завтрашний день недели
                                                            #print('[FIND] Ячейка:', sheet['A' + str(number_day_next)].value)
                                                            
                                                            #фикс бага с воскресеньем
                                                            if day_next == 'Воскресенье': #если след. день недели воскресенье, парсим плюс 10 ячеек вниз
                                                                number_day_next = number_day + 9
                                                                
                                                            list_object = range(number_day, number_day_next) #от сегодняшнего дня недели до завтрашнего
                                                            name = []
                                                            time = []
                                                            classroom = []
                                                            for obj in list_object:
                                                                if type(sheet[key + str(obj)].value) != type(None):
                                                                    if type(key) != type(None):
                                                                        classroom = a_dict.get(key)+1

                                                                    s = sheet[key + str(obj)].value + str(' | ауд. ' + str(sheet[a_dict1.get(classroom) + str(obj)].value) if type(sheet[a_dict1.get(classroom) + str(obj)].value) != type(None) else '')# найденные ячейки с парами



                                                                    patter = r'[А-ЯЁа-яё]+\s[А-ЯЁа-яё]{1}\.\s*[А-ЯЁа-яё]{1}\.\s*\/*\s*[А-ЯЁа-яё]*\s*[А-ЯЁа-яё]*\.*\s*[А-ЯЁа-яё]*\.*' # ищем фамилии
                                                                    name.append(re.findall(patter, s)) # заносим найденные фамилии в ячейках s в список name
                                                                    if 'Вакансия' in s:
                                                                        name.append([s])


                                                                    #ВРЕМЯ | TIME
                                                                    patterTIME = r'[0-9][0-9]\:[0-9][0-9][ \t]*\-*\—*[ \t]*[0-9][0-9]\:[0-9][0-9]\,*[0-9]*[0-9]*\:*[0-9]*[0-9]*[ \t]*\-*\—*[ \t]*[0-9]*[0-9]*\:*[0-9]*[0-9]*[ \t]*\-*\—*[ \t]*[0-9]*[0-9]*\:*[0-9]*[0-9]*'
                                                                    time.append(re.findall(patterTIME, s))

                                                                    #ЗАПОЛНЯЕМ СПИСОК РАСПИСАНИЕМ | weekly_schedule
                                                                    weekly_schedule.append(s) # заполянем список парами

                                                            #ФОРМАТИРУМ МАТРИЦУ С ФАМИЛИЯМИ
                                                            name = ([x for x in name if x])
                                                            name = [el for el, _ in groupby(name)]
                                                            name = sum(name, [])

                                                            name = [line.rstrip() for line in name] #удаляем символы \n из строки фамилий

                                                            print('NAME:', name)

                                                            #ФОРМАТИРУМ МАТРИЦУ С ВРЕМЕНЕМ
                                                            time = ([x for x in time if x])
                                                            time = [el for el, _ in groupby(time)]
                                                            time = sum(time, [])

                                                            time = [line.rstrip() for line in time] #удаляем символы \n из строки времени

                                                            #print('TIME:', time, 'Пара начинается в', time[0][:5], 'Домой в', time[-1][-5:])

                                                            result = []
                                                            tmp = []
                                                            for i in range(len(weekly_schedule)):
                                                                if weekly_schedule[i] in name:
                                                                    tmp.append(weekly_schedule[i - 1] + '\n' + weekly_schedule[i])
                                                                elif i + 1 < len(weekly_schedule) and weekly_schedule[i + 1] in name:
                                                                    tmp.append('')
                                                                else:
                                                                    tmp.append(weekly_schedule[i])
                                                            result = [x for x in tmp if x != '']

                                                            
                                                            #result = [w for w in weekly_schedule if w not in name else: w + name[name.index(w)]]
                                                            print(result)

                                                            """
                                                            time_current = []
                                                            for i, (start, end) in enumerate(zip(time[::2], time[1::2]), 1):
                                                                time_current.append(f"{start} - {end}")
                                                                #   if re.search(patterTIME, s) is not None:

                                                            print(time_current)
                                                            """

                                                            pTIME = r'[0-9][0-9]\:[0-9][0-9]'

                                                            count = 0
                                                            for item_i in range(len(result)):
                                                                if re.search(pTIME, result[item_i]) is not None:
                                                                    patterTIME_del = r'[0-9][0-9]\:[0-9][0-9][ \t]*\-*\—*[ \t]*[0-9][0-9]\:[0-9][0-9]\,*[0-9]*[0-9]*\:*[0-9]*[0-9]*[ \t]*\-*\—*[ \t]*[0-9]*[0-9]*\:*[0-9]*[0-9]*[ \t]*\-*\—*[ \t]*[0-9]*[0-9]*\:*[0-9]*[0-9]*'
                                                                    result[item_i] = re.sub(patterTIME_del, '', result[item_i]) # удаляем лишнее время из списка (то, что без скобок)

                                                                    result[item_i] = eval('"' + result[item_i].replace('\n','') + '"') #удалять лишнии переносы

                                                                    #print('COUNT:', count, ' item_i:', item_i)
                                                                    result[item_i] += '  (' + time[count] + ')' #добавляем скобки
                                                                    count += 1

                                                            result = [line.rstrip() for line in result] #удаляем символы \n из строки результата

                                                            
                                                            result_time = copy.deepcopy(result) #создаем копию списка расписания
                                                            print('Первая пара, номер',result_time[0][0])
                                                            print('Вторая пара, номер',result_time[1][0])
                                                            for i in range(0,5):
                                                                if result_time[i][3:] != 'Н/Б':
                                                                    time_one = str(result_time[i][0])[0]
                                                                    break
                                                            #time_one = str(result_time[0][0])[0]
                                                            time_one = str(time_schedule1.get(int("".join(time_one.split()))))[:5]
                                                            #узнаем время
                                                            for item in result_time:
                                                                if result_time[result_time.index(item)][3:6] != 'Н/Б':
                                                                    print(result_time[result_time.index(item)])
                                                                    result_time[result_time.index(item)] = ' ('+time_schedule1.get(int(item[:1]))+')'
                                                                else:
                                                                    result_time[result_time.index(item)] = ''
                                                             
                                                            print(result_time)
                                                             #= [item[1] for item in result_time]
                                                            print('В колледж в', time_one)
                                                            
                                                            
                                                            result2 = result
                                                            
                                                            result = [f'\n {j}' for j in result] # добавляем переносы перед каждым пунктом
                                                            
                                                            #заполняем временем
                                                            if len(result) == len(result_time):
                                                                result = [val for pair in zip(result, result_time) for val in pair]
                                                            else:
                                                                return 'Размер расписания не соответсутет размеру времени'

                                                            result = [f'\n\n&#128341;{to_day_or_not_today.title()} в ' + time_one] + result
                                                            result = ['Расписание на' + to_day_or_not_today] + result
                                                            

                                                            
                                                            if not merge:
                                                                print('\nРЕЗУЛЬТАТ:', result)

                                                                return (result)
                                                            elif merge:
                                                                print('\nРЕЗУЛЬТАТ2:', result2)

                                                                return result2








    #print(weekly_schedule) # выводим результат


def find_last_time(group, sheet):
    liva_weekday = weekday_list1.get(datetime.datetime.today().isoweekday())
    if liva_weekday == 'Воскресенье':
        liva_weekday = 'Понедельник'
    #liva_weekday = 'Среда'
    weekly_schedule = []
    
    
    to_day_or_not_today = ''

    #текущее время
    offset = datetime.timezone(datetime.timedelta(hours=3))
    now = datetime.datetime.now(offset)
    now = now.strftime("%H:%M")

    print('Текущее время:', now)




    


    #liva_weekday = 'Пятница'
    print('Будет показано расписание на: ', liva_weekday)
    
    print('Группа:', group, 'День недели', liva_weekday)
    for number in a: #цифры/строки
        for key in a_dict: #буквы/столбцы
            if type(sheet[key + str(number)].value) != type(None): #если ячейка не пустая
                if group.upper() in sheet[key + str(number)].value: #ищем группу в расписании
                    #пробегаем в цикле по ячейки для поиска дня недели
                    for day in weekday_list: # пробигаемся по дням неделям
                        if day == liva_weekday: # если мы нашли день недели равным текущему
                            for number_day in weekday_range: # пробигаемся по определенном диапозоне для поиска ...
                                #if type(sheet[key + str(number_day)].value) != type(None) or type(sheet['A' + str(number_day)].value) != type(None): #если ячейка не пустая
                                if type(sheet['A' + str(number_day)].value) != type(None): #если ячейка не пустая
                                
                                    if (sheet['A' + str(number_day)].value).find(day.upper()) != -1: #Нашли день недели
                                        print('[FIND] Ячейка:', sheet['A' + str(number_day)].value)
                                        next_day = weekday_list.get(day)+1 # поменчаем, что неделя следующая (если сейчас 5, то парсим до 6)
                                        lest_day_text = weekday_list1.get(next_day)# ищем номер недели в списке (чтобы писалась неделя текстом, а не цифрой)
                                        
                                        next_weekday_range = range(number_day, 100) # Диапозон от сегодняшнего дня недели до не найденного следующего (будем искать от сегодняшней до 100 ячейки)
                                        
                                        for day_next in weekday_list:
                                            if day_next == lest_day_text:
                                                for number_day_next in next_weekday_range: #пробегаем в цикле по ячейки для поиска след дня недели
                                                    #if type(sheet[key + str(number_day_next)].value) != type(None) or type(sheet['A' + str(number_day_next)].value) != type(None): #если ячейка не пустая
                                                    if type(sheet['A' + str(number_day_next)].value) != type(None): #если ячейка не пустая
                                                    
                                                        if (sheet['A' + str(number_day_next)].value).find(day_next.upper()) != -1 or day_next == 'Воскресенье': #Нашли завтрашний день недели
                                                            #print('[FIND] Ячейка:', sheet['A' + str(number_day_next)].value)
                                                            
                                                            if day_next == 'Воскресенье': #если след. день недели воскресенье, парсим плюс 10 ячеек вниз
                                                                number_day_next = number_day + 9
                                                            
                                                            list_object = range(number_day, number_day_next) #от сегодняшнего дня недели до завтрашнего
                                                            name = []
                                                            time = []
                                                            for obj in list_object:
                                                                if type(sheet[key + str(obj)].value) != type(None):

                                                                    #ФАМИЛИИ | NAME
                                                                    s = sheet[key + str(obj)].value # найденные ячейки с парами
                                                                    patter = r'[А-ЯЁа-яё]+\s[А-ЯЁа-яё]{1}\.\s*[А-ЯЁа-яё]{1}\.\s*\/*\s*[А-ЯЁа-яё]*\s*[А-ЯЁа-яё]*\.*\s*[А-ЯЁа-яё]*\.*' # ищем фамилии
                                                                    name.append(re.findall(patter, s)) # заносим найденные фамилии в ячейках s в список name
                                                                    
                                                                    if 'Вакансия' in s:
                                                                        name.append([s])
                                                                    

                                                                    #ЗАПОЛНЯЕМ СПИСОК РАСПИСАНИЕМ | weekly_schedule
                                                                    weekly_schedule.append(s) # заполянем список парами

                                                            #ФОРМАТИРУМ МАТРИЦУ С ФАМИЛИЯМИ
                                                            name = ([x for x in name if x])
                                                            name = [el for el, _ in groupby(name)]
                                                            name = sum(name, [])

                                                            name = [line.rstrip() for line in name] #удаляем символы \n из строки фамилий

                                                            print('NAME:', name)

                                                            

                                                            result = [w for w in weekly_schedule if w not in name]

                                                            """
                                                            time_current = []
                                                            for i, (start, end) in enumerate(zip(time[::2], time[1::2]), 1):
                                                                time_current.append(f"{start} - {end}")
                                                                #   if re.search(patterTIME, s) is not None:

                                                            print(time_current)
                                                            """

                                                            

                                                            count = 0
                                                            

                                                            result = [line.rstrip() for line in result] #удаляем символы \n из строки результата

                                                            #print(result)

                                                            for item in result:
                                                                result[result.index(item)] = item[:1]
                                                            
                                                            #print(str(result).replace(' ', ''))
                                                            last_para = time_schedule1.get(int(str(result[-1]).replace(' ', '')))
                                                            print('\nРЕЗУЛЬТАТ:', last_para[8:])
                                                            return last_para[8:]

#main_open('4исп-9', 'Расписание_занятий_2-4_курсов_с_13.09.21_(2_поток)', 'Расписание', False, False)
#main_open('2зио-9', 'Расписание занятий 1-2 курсов с 13.09.21 (1 поток)','Узнать на какой день недели парсить замены', False, False)